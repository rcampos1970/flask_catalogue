from flask import Flask, render_template, request
from openpyxl import load_workbook
app = Flask(__name__)


@app.route("/", methods=["POST", "GET"])
def home():
    while True:
        print(request.form["nm"])
        print(request.form["ds"])
        print(request.form["pr"])
        name = request.form["nm"]
        description = request.form["ds"]
        price = request.form["pr"]
        wb = load_workbook("Cathalogue.xlsx")
        ws = wb.active
        column = ws.cell(7, 19).value
        print("column is", column)
        ws.cell(column, 1, name)
        ws.cell(column, 2, description)
        ws.cell(column, 3, price)
        column = column + 1
        ws.cell(7, 19, column)
        wb.save("Cathalogue.xlsx")
        return render_template("index.html")


if __name__ == "__main__":
    app.run()
